"""
Restore de dados da Pesquisa G4 a partir de Excels exportados.

Re-cria áreas, funcionarios, tokens e respostas (incluindo nps_areas e
avaliacoes_individuais) a partir de 2 arquivos:
  - Tokens Excel (exportado de /admin/exportar-tokens) → tem pessoa + email + token
  - Resultados Excel (exportado de /admin/exportar-resultados) → tem áreas + respostas

Funcionarios vem do Excel de tokens (são os 484 que receberam link).
Áreas vem dos resultados (31 áreas já com os splits customizados).
"""

from collections import defaultdict
from openpyxl import load_workbook


RESTORE_FORM_HTML = """
<!DOCTYPE html>
<html lang="pt-BR">
<head>
<meta charset="UTF-8">
<title>Restore — Pesquisa G4</title>
<link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.2/dist/css/bootstrap.min.css" rel="stylesheet">
<style>
  body { background: #f5f6f8; font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif; padding: 30px; }
  .container { max-width: 720px; }
  .card { border: none; border-radius: 12px; box-shadow: 0 2px 8px rgba(0,0,0,0.06); }
</style>
</head>
<body>
<div class="container">
  <h1 class="mb-3">Restore da Base</h1>
  <p class="text-muted">Reconstrói o SQLite a partir dos Excels exportados. <strong>Atenção:</strong> APAGA os dados atuais antes de reinserir.</p>

  {% if result %}
    <div class="alert {% if result.ok %}alert-success{% else %}alert-danger{% endif %}">
      <h5>{% if result.ok %}✓ Restore concluído{% else %}✗ Erro no restore{% endif %}</h5>
      <pre class="mb-0 small">{{ result.msg }}</pre>
    </div>
  {% endif %}

  <div class="card p-4">
    <form method="POST" action="/admin/restore" enctype="multipart/form-data">
      <div class="mb-3">
        <label class="form-label"><strong>1. Tokens Excel</strong> <small class="text-muted">(tokens_pesquisa*.xlsx — provê funcionários + tokens)</small></label>
        <input type="file" class="form-control" name="tokens" accept=".xlsx" required>
      </div>

      <div class="mb-3">
        <label class="form-label"><strong>2. Resultados Excel</strong> <small class="text-muted">(resultados_pesquisa_*.xlsx — provê áreas + respostas)</small></label>
        <input type="file" class="form-control" name="resultados" accept=".xlsx" required>
      </div>

      <div class="form-check mb-3">
        <input class="form-check-input" type="checkbox" id="confirm" name="confirm" value="1" required>
        <label class="form-check-label" for="confirm">
          Confirmo que quero APAGAR os dados atuais e reconstruir a partir destes arquivos
        </label>
      </div>

      <button type="submit" class="btn btn-danger">Executar Restore</button>
      <a href="/admin" class="btn btn-outline-secondary ms-2">Cancelar</a>
    </form>
  </div>
</div>
</body>
</html>
"""


def load_tokens(path):
    """Le o Excel de tokens exportados. Retorna lista de dicts com token + pessoa."""
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    tokens = []
    for r in rows[1:]:
        if not r or len(r) < 4:
            continue
        pessoa = r[1]
        email = r[2]
        link = r[3]
        if not link:
            continue
        token_str = str(link).rstrip('/').split('/')[-1]
        if not token_str:
            continue
        tokens.append({
            'token': token_str,
            'pessoa_nome': str(pessoa).strip().upper() if pessoa else None,
            'pessoa_email': str(email).strip().lower() if email else None,
        })
    return tokens


def load_resultados(path):
    """Le o Excel de resultados e retorna (areas, respostas, nps_rows, ind_rows)."""
    wb = load_workbook(path, data_only=True)

    # Areas — da aba NPS por Area (primeira coluna)
    areas_set = set()
    ws_areas = wb['NPS por Area']
    for r in list(ws_areas.iter_rows(values_only=True))[1:]:
        if r and r[0]:
            areas_set.add(str(r[0]).strip())

    # Respostas + NPS — da aba Respostas NPS Detalhadas
    ws_nps = wb['Respostas NPS Detalhadas']
    nps_raw = list(ws_nps.iter_rows(values_only=True))[1:]

    # Individuais — da aba Individuais Detalhado
    ws_ind = wb['Individuais Detalhado']
    ind_raw = list(ws_ind.iter_rows(values_only=True))[1:]

    # Adicionar areas que aparecem como respondente também
    for r in nps_raw:
        if r and r[1]:
            areas_set.add(str(r[1]).strip())

    respostas = {}
    nps_por_resposta = defaultdict(list)
    for r in nps_raw:
        if not r or r[0] is None:
            continue
        resp_id, area_resp, senior, ano, area_aval, nota, coment, data = r
        if resp_id not in respostas:
            respostas[resp_id] = {
                'area_respondente': str(area_resp).strip() if area_resp else None,
                'senioridade': str(senior).strip() if senior else None,
                'ano_entrada': str(ano).strip() if ano is not None else None,
                'criado_em': str(data) if data else None,
            }
        nps_por_resposta[resp_id].append({
            'area_avaliada': str(area_aval).strip() if area_aval else None,
            'nota': int(nota) if nota is not None else None,
            'comentario': coment,
        })

    ind_por_resposta = defaultdict(list)
    for r in ind_raw:
        if not r or r[0] is None:
            continue
        resp_id, area_resp, senior, ano, func, recomenda, motivo, data = r
        if resp_id not in respostas:
            respostas[resp_id] = {
                'area_respondente': str(area_resp).strip() if area_resp else None,
                'senioridade': str(senior).strip() if senior else None,
                'ano_entrada': str(ano).strip() if ano is not None else None,
                'criado_em': str(data) if data else None,
            }
        recomenda_str = str(recomenda).strip().lower() if recomenda else ''
        recomenda_int = 1 if recomenda_str in ('sim', 'yes', '1', 'true') else 0
        ind_por_resposta[resp_id].append({
            'funcionario_avaliado': str(func).strip() if func else None,
            'recomenda': recomenda_int,
            'motivo': motivo,
        })

    return sorted(areas_set), respostas, nps_por_resposta, ind_por_resposta


def executar_restore(conn, tokens_path, resultados_path):
    """Apaga dados atuais e reinsere a partir dos 2 Excels. Retorna resumo (dict)."""
    tokens = load_tokens(tokens_path)
    areas, respostas, nps_por_r, ind_por_r = load_resultados(resultados_path)

    # Funcionarios vem dos tokens (deduplicar por nome)
    funcionarios_seen = {}
    for t in tokens:
        nome = t.get('pessoa_nome')
        if nome and nome not in funcionarios_seen:
            funcionarios_seen[nome] = t.get('pessoa_email')
    funcionarios = [{'nome': n, 'email': e} for n, e in funcionarios_seen.items()]

    # Limpar tabelas (ordem importa por FK)
    conn.execute('DELETE FROM avaliacoes_individuais')
    conn.execute('DELETE FROM nps_areas')
    conn.execute('DELETE FROM respostas')
    conn.execute('DELETE FROM tokens')
    conn.execute('DELETE FROM funcionarios')
    conn.execute('DELETE FROM areas')

    # Inserir areas
    for nome in areas:
        conn.execute('INSERT INTO areas (nome) VALUES (?)', (nome,))

    # Inserir funcionarios (area_id=NULL porque nao temos mapping confiavel)
    for f in funcionarios:
        conn.execute(
            'INSERT INTO funcionarios (nome, area_id, email) VALUES (?, NULL, ?)',
            (f['nome'], f.get('email'))
        )

    # Inserir tokens (todos como nao usados)
    for t in tokens:
        conn.execute(
            'INSERT INTO tokens (token, usado, pessoa_nome, pessoa_email) VALUES (?, 0, ?, ?)',
            (t['token'], t.get('pessoa_nome'), t.get('pessoa_email'))
        )

    # Inserir respostas + dependentes
    nps_inseridos = 0
    ind_inseridos = 0
    for resp_id, meta in respostas.items():
        cur = conn.execute(
            'INSERT INTO respostas (area_respondente, senioridade, ano_entrada, criado_em) VALUES (?, ?, ?, ?)',
            (meta['area_respondente'], meta['senioridade'], meta.get('ano_entrada'), meta.get('criado_em'))
        )
        new_id = cur.lastrowid

        for n in nps_por_r.get(resp_id, []):
            conn.execute(
                'INSERT INTO nps_areas (resposta_id, area_avaliada, nota, comentario) VALUES (?, ?, ?, ?)',
                (new_id, n['area_avaliada'], n['nota'], n.get('comentario'))
            )
            nps_inseridos += 1

        for i in ind_por_r.get(resp_id, []):
            conn.execute(
                'INSERT INTO avaliacoes_individuais (resposta_id, funcionario_avaliado, recomenda, motivo) VALUES (?, ?, ?, ?)',
                (new_id, i['funcionario_avaliado'], i['recomenda'], i.get('motivo'))
            )
            ind_inseridos += 1

    conn.commit()

    return {
        'areas': len(areas),
        'funcionarios': len(funcionarios),
        'tokens': len(tokens),
        'respostas': len(respostas),
        'nps_votes': nps_inseridos,
        'avaliacoes_individuais': ind_inseridos,
    }

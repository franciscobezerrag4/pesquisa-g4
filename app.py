"""
Pesquisa de Qualidade das Entregas - G4
Sistema de survey anonimo com controle de resposta unica.
Deploy: Railway (HTTPS automatico)
"""

import os
import json
import sqlite3
import secrets
from datetime import datetime
from functools import wraps
from flask import Flask, render_template, request, jsonify, redirect, url_for, send_file, session, Response
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', secrets.token_hex(32))

# Config
ADMIN_USER = os.environ.get('ADMIN_USER', 'admin')
ADMIN_PASS = os.environ.get('ADMIN_PASS', 'g4pesquisa2026')

DATA_DIR = os.environ.get('DATA_DIR', os.path.dirname(__file__))
DB_PATH = os.path.join(DATA_DIR, 'pesquisa.db')
UPLOAD_FOLDER = os.path.join(DATA_DIR, 'uploads')
EXPORT_FOLDER = os.path.join(DATA_DIR, 'exports')

os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(EXPORT_FOLDER, exist_ok=True)


# === Auth ===
def check_auth(username, password):
    return username == ADMIN_USER and password == ADMIN_PASS


def auth_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        auth = request.authorization
        if not auth or not check_auth(auth.username, auth.password):
            return Response(
                'Acesso negado. Credenciais invalidas.',
                401,
                {'WWW-Authenticate': 'Basic realm="Admin Pesquisa G4"'}
            )
        return f(*args, **kwargs)
    return decorated


# === DB ===
def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def init_db():
    conn = get_db()
    conn.executescript('''
        CREATE TABLE IF NOT EXISTS areas (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nome TEXT UNIQUE NOT NULL
        );
        CREATE TABLE IF NOT EXISTS funcionarios (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nome TEXT NOT NULL,
            email TEXT,
            area_id INTEGER,
            FOREIGN KEY (area_id) REFERENCES areas(id)
        );
        CREATE TABLE IF NOT EXISTS tokens (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            token TEXT UNIQUE NOT NULL,
            usado INTEGER DEFAULT 0,
            criado_em TEXT DEFAULT CURRENT_TIMESTAMP,
            usado_em TEXT
        );
        CREATE TABLE IF NOT EXISTS respostas (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            token_id INTEGER NOT NULL,
            area_respondente TEXT NOT NULL,
            senioridade TEXT NOT NULL,
            ano_entrada TEXT,
            criado_em TEXT DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (token_id) REFERENCES tokens(id)
        );
        CREATE TABLE IF NOT EXISTS nps_areas (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            resposta_id INTEGER NOT NULL,
            area_avaliada TEXT NOT NULL,
            nota INTEGER NOT NULL,
            comentario TEXT,
            FOREIGN KEY (resposta_id) REFERENCES respostas(id)
        );
        CREATE TABLE IF NOT EXISTS avaliacoes_individuais (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            resposta_id INTEGER NOT NULL,
            funcionario_avaliado TEXT NOT NULL,
            recomenda INTEGER NOT NULL,
            motivo TEXT,
            FOREIGN KEY (resposta_id) REFERENCES respostas(id)
        );
    ''')
    conn.commit()
    conn.close()


init_db()


# Migrations: add columns if not exist
def migrate_db():
    conn = get_db()
    try:
        conn.execute('SELECT ano_entrada FROM respostas LIMIT 1')
    except sqlite3.OperationalError:
        conn.execute('ALTER TABLE respostas ADD COLUMN ano_entrada TEXT')
        conn.commit()
    try:
        conn.execute('SELECT pessoa_nome FROM tokens LIMIT 1')
    except sqlite3.OperationalError:
        conn.execute('ALTER TABLE tokens ADD COLUMN pessoa_nome TEXT')
        conn.commit()
    try:
        conn.execute('SELECT email FROM funcionarios LIMIT 1')
    except sqlite3.OperationalError:
        conn.execute('ALTER TABLE funcionarios ADD COLUMN email TEXT')
        conn.commit()
    try:
        conn.execute('SELECT pessoa_email FROM tokens LIMIT 1')
    except sqlite3.OperationalError:
        conn.execute('ALTER TABLE tokens ADD COLUMN pessoa_email TEXT')
        conn.commit()
    conn.close()


migrate_db()


# === ROTAS PUBLICAS (pesquisa) ===

@app.route('/')
def home():
    return render_template('home.html')


@app.route('/pesquisa/<token>')
def pesquisa(token):
    conn = get_db()
    t = conn.execute('SELECT id, usado FROM tokens WHERE token = ?', (token,)).fetchone()

    if not t:
        conn.close()
        return render_template('erro.html', msg='Link inválido. Verifique o link recebido.')

    if t['usado']:
        conn.close()
        return render_template('erro.html', msg='Este link já foi utilizado. Cada pessoa pode responder apenas uma vez.')

    areas = [r['nome'] for r in conn.execute('SELECT nome FROM areas ORDER BY nome').fetchall()]
    funcs = [r['nome'] for r in conn.execute('SELECT nome FROM funcionarios ORDER BY nome').fetchall()]
    conn.close()

    return render_template('pesquisa.html', token=token, areas=areas, funcionarios=funcs)


@app.route('/pesquisa/<token>/enviar', methods=['POST'])
def enviar_resposta(token):
    conn = get_db()
    t = conn.execute('SELECT id, usado FROM tokens WHERE token = ?', (token,)).fetchone()

    if not t:
        conn.close()
        return jsonify({'error': 'Token inválido'}), 400

    if t['usado']:
        conn.close()
        return jsonify({'error': 'Este link já foi utilizado'}), 400

    data = request.get_json()

    if not data.get('area_respondente') or not data.get('senioridade'):
        return jsonify({'error': 'Área e nível são obrigatórios'}), 400

    if not data.get('ano_entrada'):
        return jsonify({'error': 'Ano de entrada é obrigatório'}), 400

    if not data.get('nps_areas') or len(data['nps_areas']) != 5:
        return jsonify({'error': 'Selecione exatamente 5 áreas e avalie todas'}), 400

    if not data.get('avaliacoes') or len(data['avaliacoes']) != 10:
        return jsonify({'error': 'Selecione exatamente 10 pessoas para avaliar'}), 400

    # Valida que notas sao inteiros 0-10
    for nps in data['nps_areas']:
        nota = nps.get('nota')
        if not isinstance(nota, int) or nota < 0 or nota > 10:
            return jsonify({'error': 'Notas devem ser entre 0 e 10'}), 400

    cursor = conn.execute(
        'INSERT INTO respostas (token_id, area_respondente, senioridade, ano_entrada) VALUES (?, ?, ?, ?)',
        (t['id'], data['area_respondente'], data['senioridade'], data.get('ano_entrada', ''))
    )
    resposta_id = cursor.lastrowid

    for nps in data['nps_areas']:
        conn.execute(
            'INSERT INTO nps_areas (resposta_id, area_avaliada, nota, comentario) VALUES (?, ?, ?, ?)',
            (resposta_id, nps['area'], nps['nota'], nps.get('comentario', ''))
        )

    for av in data['avaliacoes']:
        conn.execute(
            'INSERT INTO avaliacoes_individuais (resposta_id, funcionario_avaliado, recomenda, motivo) VALUES (?, ?, ?, ?)',
            (resposta_id, av['nome'], 1 if av['recomenda'] else 0, av.get('motivo', ''))
        )

    conn.execute('UPDATE tokens SET usado = 1, usado_em = ? WHERE id = ?',
                 (datetime.now().isoformat(), t['id']))

    conn.commit()
    conn.close()

    return jsonify({'success': True})


# === ROTAS ADMIN (protegidas) ===

@app.route('/admin')
@auth_required
def admin():
    return render_template('admin.html')


@app.route('/admin/upload', methods=['POST'])
@auth_required
def upload_excel():
    if 'file' not in request.files:
        return jsonify({'error': 'Nenhum arquivo enviado'}), 400

    file = request.files['file']
    if not file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({'error': 'Arquivo deve ser .xlsx'}), 400

    filepath = os.path.join(UPLOAD_FOLDER, 'dados.xlsx')
    file.save(filepath)

    wb = load_workbook(filepath)

    conn = get_db()
    conn.execute('DELETE FROM funcionarios')
    conn.execute('DELETE FROM areas')

    if 'Areas' in wb.sheetnames:
        ws = wb['Areas']
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:
                conn.execute('INSERT OR IGNORE INTO areas (nome) VALUES (?)', (str(row[0]).strip(),))

    if 'Funcionarios' in wb.sheetnames:
        ws = wb['Funcionarios']
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:
                nome = str(row[0]).strip()
                area_nome = str(row[1]).strip() if len(row) > 1 and row[1] else None
                email = str(row[2]).strip() if len(row) > 2 and row[2] else None
                area_id = None
                if area_nome:
                    r = conn.execute('SELECT id FROM areas WHERE nome = ?', (area_nome,)).fetchone()
                    if r:
                        area_id = r['id']
                conn.execute('INSERT INTO funcionarios (nome, email, area_id) VALUES (?, ?, ?)', (nome, email, area_id))

    conn.commit()

    areas_count = conn.execute('SELECT COUNT(*) as c FROM areas').fetchone()['c']
    func_count = conn.execute('SELECT COUNT(*) as c FROM funcionarios').fetchone()['c']
    conn.close()

    return jsonify({'success': True, 'areas': areas_count, 'funcionarios': func_count})


@app.route('/admin/gerar-tokens', methods=['POST'])
@auth_required
def gerar_tokens():
    data = request.get_json()
    quantidade = int(data.get('quantidade', 0))

    if quantidade <= 0 or quantidade > 500:
        return jsonify({'error': 'Quantidade inválida (1-500)'}), 400

    conn = get_db()
    tokens = []
    for _ in range(quantidade):
        token = secrets.token_urlsafe(16)
        conn.execute('INSERT INTO tokens (token) VALUES (?)', (token,))
        tokens.append(token)
    conn.commit()
    conn.close()

    base_url = request.host_url.rstrip('/')
    links = [f"{base_url}/pesquisa/{t}" for t in tokens]

    return jsonify({'success': True, 'tokens': tokens, 'links': links})


@app.route('/admin/gerar-tokens-pessoas', methods=['POST'])
@auth_required
def gerar_tokens_pessoas():
    conn = get_db()
    funcs = conn.execute('SELECT nome, email FROM funcionarios ORDER BY nome').fetchall()

    if not funcs:
        conn.close()
        return jsonify({'error': 'Nenhum funcionário cadastrado. Faça o upload da base primeiro.'}), 400

    criados = []
    for f in funcs:
        nome = f['nome']
        email = f['email']
        existente = conn.execute(
            'SELECT token FROM tokens WHERE pessoa_nome = ?', (nome,)
        ).fetchone()
        if existente:
            continue
        token = secrets.token_urlsafe(16)
        conn.execute(
            'INSERT INTO tokens (token, pessoa_nome, pessoa_email) VALUES (?, ?, ?)',
            (token, nome, email)
        )
        criados.append({'nome': nome, 'email': email, 'token': token})

    conn.commit()
    conn.close()

    base_url = request.host_url.rstrip('/')
    result = [
        {'nome': c['nome'], 'email': c['email'] or '', 'link': f"{base_url}/pesquisa/{c['token']}"}
        for c in criados
    ]

    return jsonify({
        'success': True,
        'criados': len(criados),
        'total_funcionarios': len(funcs),
        'pessoas': result
    })


@app.route('/admin/exportar-tokens')
@auth_required
def exportar_tokens():
    conn = get_db()
    tokens = conn.execute(
        'SELECT token, pessoa_nome, pessoa_email, usado, criado_em FROM tokens ORDER BY pessoa_nome, id'
    ).fetchall()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = 'Tokens'

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')

    headers = ['#', 'Pessoa', 'Email', 'Link da Pesquisa', 'Status', 'Criado em']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    base_url = request.host_url.rstrip('/')
    for i, t in enumerate(tokens, 1):
        ws.cell(row=i+1, column=1, value=i)
        ws.cell(row=i+1, column=2, value=t['pessoa_nome'] or '(anônimo)')
        ws.cell(row=i+1, column=3, value=t['pessoa_email'] or '')
        ws.cell(row=i+1, column=4, value=f"{base_url}/pesquisa/{t['token']}")
        ws.cell(row=i+1, column=5, value='Respondido' if t['usado'] else 'Pendente')
        ws.cell(row=i+1, column=6, value=t['criado_em'])

    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 60
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 20

    filepath = os.path.join(EXPORT_FOLDER, 'tokens_pesquisa.xlsx')
    wb.save(filepath)
    return send_file(filepath, as_attachment=True)


@app.route('/admin/dados')
@auth_required
def admin_dados():
    conn = get_db()
    areas = [r['nome'] for r in conn.execute('SELECT nome FROM areas ORDER BY nome').fetchall()]
    funcs = conn.execute('''
        SELECT f.nome, COALESCE(a.nome, '') as area
        FROM funcionarios f LEFT JOIN areas a ON f.area_id = a.id
        ORDER BY f.nome
    ''').fetchall()
    funcionarios = [{'nome': f['nome'], 'area': f['area']} for f in funcs]

    tokens_total = conn.execute('SELECT COUNT(*) as c FROM tokens').fetchone()['c']
    tokens_usados = conn.execute('SELECT COUNT(*) as c FROM tokens WHERE usado = 1').fetchone()['c']
    respostas_total = conn.execute('SELECT COUNT(*) as c FROM respostas').fetchone()['c']

    conn.close()

    return jsonify({
        'areas': areas,
        'funcionarios': funcionarios,
        'tokens_total': tokens_total,
        'tokens_usados': tokens_usados,
        'respostas_total': respostas_total
    })


@app.route('/admin/exportar-resultados')
@auth_required
def exportar_resultados():
    conn = get_db()

    wb = Workbook()
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    def style_header(ws, headers):
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
            cell.border = thin_border

    # Aba 1: Resumo NPS por Area
    ws1 = wb.active
    ws1.title = 'NPS por Area'
    headers1 = ['Area Avaliada', 'Total Avaliacoes', 'Media', 'Promotores (9-10)', 'Neutros (7-8)',
                 'Detratores (0-6)', '% Promotores', '% Detratores', 'NPS Score']
    style_header(ws1, headers1)

    areas_nps = conn.execute('''
        SELECT area_avaliada, COUNT(*) as total,
               AVG(nota) as media,
               SUM(CASE WHEN nota >= 9 THEN 1 ELSE 0 END) as promotores,
               SUM(CASE WHEN nota >= 7 AND nota <= 8 THEN 1 ELSE 0 END) as neutros,
               SUM(CASE WHEN nota <= 6 THEN 1 ELSE 0 END) as detratores
        FROM nps_areas
        GROUP BY area_avaliada
        ORDER BY area_avaliada
    ''').fetchall()

    for i, r in enumerate(areas_nps, 2):
        total = r['total']
        ws1.cell(row=i, column=1, value=r['area_avaliada'])
        ws1.cell(row=i, column=2, value=total)
        ws1.cell(row=i, column=3, value=round(r['media'], 2))
        ws1.cell(row=i, column=4, value=r['promotores'])
        ws1.cell(row=i, column=5, value=r['neutros'])
        ws1.cell(row=i, column=6, value=r['detratores'])
        pct_prom = round(r['promotores'] / total * 100, 1) if total else 0
        pct_det = round(r['detratores'] / total * 100, 1) if total else 0
        ws1.cell(row=i, column=7, value=pct_prom)
        ws1.cell(row=i, column=8, value=pct_det)
        ws1.cell(row=i, column=9, value=round(pct_prom - pct_det, 1))

    for col in 'ABCDEFGHI':
        ws1.column_dimensions[col].width = 18

    # Aba 2: NPS Detalhado
    ws2 = wb.create_sheet('Respostas NPS Detalhadas')
    headers2 = ['Resposta #', 'Area do Respondente', 'Senioridade', 'Ano de Entrada', 'Area Avaliada', 'Nota', 'Comentario', 'Data']
    style_header(ws2, headers2)

    detalhes = conn.execute('''
        SELECT r.id, r.area_respondente, r.senioridade, r.ano_entrada, n.area_avaliada, n.nota, n.comentario, r.criado_em
        FROM respostas r
        JOIN nps_areas n ON n.resposta_id = r.id
        ORDER BY r.id, n.area_avaliada
    ''').fetchall()

    for i, d in enumerate(detalhes, 2):
        ws2.cell(row=i, column=1, value=d['id'])
        ws2.cell(row=i, column=2, value=d['area_respondente'])
        ws2.cell(row=i, column=3, value=d['senioridade'])
        ws2.cell(row=i, column=4, value=d['ano_entrada'])
        ws2.cell(row=i, column=5, value=d['area_avaliada'])
        ws2.cell(row=i, column=6, value=d['nota'])
        ws2.cell(row=i, column=7, value=d['comentario'])
        ws2.cell(row=i, column=8, value=d['criado_em'])

    for col in 'ABCDEFGH':
        ws2.column_dimensions[col].width = 22

    # Aba 3: NPS por Senioridade
    ws3 = wb.create_sheet('NPS por Senioridade')
    headers3 = ['Area Avaliada', 'Senioridade', 'Total', 'Media', 'NPS Score']
    style_header(ws3, headers3)

    cross = conn.execute('''
        SELECT n.area_avaliada, r.senioridade, COUNT(*) as total,
               AVG(n.nota) as media,
               SUM(CASE WHEN n.nota >= 9 THEN 1 ELSE 0 END) as promotores,
               SUM(CASE WHEN n.nota <= 6 THEN 1 ELSE 0 END) as detratores
        FROM nps_areas n
        JOIN respostas r ON n.resposta_id = r.id
        GROUP BY n.area_avaliada, r.senioridade
        ORDER BY n.area_avaliada, r.senioridade
    ''').fetchall()

    for i, c in enumerate(cross, 2):
        total = c['total']
        pct_prom = round(c['promotores'] / total * 100, 1) if total else 0
        pct_det = round(c['detratores'] / total * 100, 1) if total else 0
        ws3.cell(row=i, column=1, value=c['area_avaliada'])
        ws3.cell(row=i, column=2, value=c['senioridade'])
        ws3.cell(row=i, column=3, value=total)
        ws3.cell(row=i, column=4, value=round(c['media'], 2))
        ws3.cell(row=i, column=5, value=round(pct_prom - pct_det, 1))

    for col in 'ABCDE':
        ws3.column_dimensions[col].width = 22

    # Aba 4: Individuais Resumo
    ws4 = wb.create_sheet('Individuais Resumo')
    headers4 = ['Funcionario', 'Total Avaliacoes', 'Recomendacoes', 'Nao Recomendacoes', '% Recomendacao']
    style_header(ws4, headers4)

    individuais = conn.execute('''
        SELECT funcionario_avaliado,
               COUNT(*) as total,
               SUM(CASE WHEN recomenda = 1 THEN 1 ELSE 0 END) as sim,
               SUM(CASE WHEN recomenda = 0 THEN 1 ELSE 0 END) as nao
        FROM avaliacoes_individuais
        GROUP BY funcionario_avaliado
        ORDER BY funcionario_avaliado
    ''').fetchall()

    for i, ind in enumerate(individuais, 2):
        total = ind['total']
        ws4.cell(row=i, column=1, value=ind['funcionario_avaliado'])
        ws4.cell(row=i, column=2, value=total)
        ws4.cell(row=i, column=3, value=ind['sim'])
        ws4.cell(row=i, column=4, value=ind['nao'])
        ws4.cell(row=i, column=5, value=round(ind['sim'] / total * 100, 1) if total else 0)

    for col in 'ABCDE':
        ws4.column_dimensions[col].width = 22

    # Aba 5: Individuais Detalhado
    ws5 = wb.create_sheet('Individuais Detalhado')
    headers5 = ['Resposta #', 'Area do Respondente', 'Senioridade', 'Ano de Entrada', 'Funcionario Avaliado', 'Recomenda', 'Motivo', 'Data']
    style_header(ws5, headers5)

    ind_det = conn.execute('''
        SELECT r.id, r.area_respondente, r.senioridade, r.ano_entrada, a.funcionario_avaliado, a.recomenda, a.motivo, r.criado_em
        FROM respostas r
        JOIN avaliacoes_individuais a ON a.resposta_id = r.id
        ORDER BY r.id, a.funcionario_avaliado
    ''').fetchall()

    for i, d in enumerate(ind_det, 2):
        ws5.cell(row=i, column=1, value=d['id'])
        ws5.cell(row=i, column=2, value=d['area_respondente'])
        ws5.cell(row=i, column=3, value=d['senioridade'])
        ws5.cell(row=i, column=4, value=d['ano_entrada'])
        ws5.cell(row=i, column=5, value=d['funcionario_avaliado'])
        ws5.cell(row=i, column=6, value='Sim' if d['recomenda'] else 'Nao')
        ws5.cell(row=i, column=7, value=d['motivo'])
        ws5.cell(row=i, column=8, value=d['criado_em'])

    for col in 'ABCDEFGH':
        ws5.column_dimensions[col].width = 22

    # Aba 6: Comentarios
    ws6 = wb.create_sheet('Comentarios')
    headers6 = ['Tipo', 'Area Respondente', 'Senioridade', 'Sobre (Area ou Pessoa)', 'Comentario/Motivo']
    style_header(ws6, headers6)

    row_num = 2
    coments_nps = conn.execute('''
        SELECT r.area_respondente, r.senioridade, n.area_avaliada, n.comentario
        FROM nps_areas n
        JOIN respostas r ON n.resposta_id = r.id
        WHERE n.comentario IS NOT NULL AND n.comentario != ''
        ORDER BY n.area_avaliada
    ''').fetchall()

    for c in coments_nps:
        ws6.cell(row=row_num, column=1, value='NPS Area')
        ws6.cell(row=row_num, column=2, value=c['area_respondente'])
        ws6.cell(row=row_num, column=3, value=c['senioridade'])
        ws6.cell(row=row_num, column=4, value=c['area_avaliada'])
        ws6.cell(row=row_num, column=5, value=c['comentario'])
        row_num += 1

    motivos = conn.execute('''
        SELECT r.area_respondente, r.senioridade, a.funcionario_avaliado, a.motivo, a.recomenda
        FROM avaliacoes_individuais a
        JOIN respostas r ON a.resposta_id = r.id
        WHERE a.motivo IS NOT NULL AND a.motivo != ''
        ORDER BY a.funcionario_avaliado
    ''').fetchall()

    for m in motivos:
        tipo = 'Individual (Recomenda)' if m['recomenda'] else 'Individual (Nao Recomenda)'
        ws6.cell(row=row_num, column=1, value=tipo)
        ws6.cell(row=row_num, column=2, value=m['area_respondente'])
        ws6.cell(row=row_num, column=3, value=m['senioridade'])
        ws6.cell(row=row_num, column=4, value=m['funcionario_avaliado'])
        ws6.cell(row=row_num, column=5, value=m['motivo'])
        row_num += 1

    for col in 'ABCDE':
        ws6.column_dimensions[col].width = 28

    conn.close()

    filepath = os.path.join(EXPORT_FOLDER, f'resultados_pesquisa_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')
    wb.save(filepath)
    return send_file(filepath, as_attachment=True)


@app.route('/admin/gerar-template')
@auth_required
def gerar_template():
    wb = Workbook()
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    example_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')

    ws1 = wb.active
    ws1.title = 'Areas'
    cell = ws1.cell(row=1, column=1, value='Nome da Area')
    cell.font = header_font
    cell.fill = header_fill
    ws1.column_dimensions['A'].width = 35

    exemplos_areas = ['Tecnologia', 'Marketing', 'Financeiro', 'RH', 'Comercial', 'Operacoes', 'Juridico']
    for i, ex in enumerate(exemplos_areas, 2):
        cell = ws1.cell(row=i, column=1, value=ex)
        cell.fill = example_fill

    ws2 = wb.create_sheet('Funcionarios')
    for col, h in enumerate(['Nome Completo', 'Area'], 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill

    exemplos_func = [
        ('Joao Silva', 'Tecnologia'),
        ('Maria Santos', 'Marketing'),
        ('Pedro Oliveira', 'Financeiro'),
    ]
    for i, (nome, area) in enumerate(exemplos_func, 2):
        ws2.cell(row=i, column=1, value=nome).fill = example_fill
        ws2.cell(row=i, column=2, value=area).fill = example_fill

    ws2.column_dimensions['A'].width = 35
    ws2.column_dimensions['B'].width = 25

    ws3 = wb.create_sheet('Instrucoes')
    instrucoes = [
        'INSTRUCOES DE PREENCHIMENTO',
        '',
        '1. Aba "Areas": Liste todas as areas da empresa (uma por linha)',
        '   - Coluna A: Nome da area',
        '   - Apague os exemplos em verde e coloque os dados reais',
        '',
        '2. Aba "Funcionarios": Liste todos os funcionarios da empresa',
        '   - Coluna A: Nome completo do funcionario',
        '   - Coluna B: Area a que pertence (deve ser igual ao nome na aba Areas)',
        '   - Apague os exemplos em verde e coloque os dados reais',
        '',
        '3. Apos preencher, salve o arquivo e faca o upload no painel admin',
        '',
        'IMPORTANTE: Os nomes das areas na aba Funcionarios devem ser',
        'exatamente iguais aos nomes na aba Areas.',
    ]
    for i, texto in enumerate(instrucoes, 1):
        cell = ws3.cell(row=i, column=1, value=texto)
        if i == 1:
            cell.font = Font(bold=True, size=14)
    ws3.column_dimensions['A'].width = 65

    filepath = os.path.join(EXPORT_FOLDER, 'template_pesquisa_g4.xlsx')
    wb.save(filepath)
    return send_file(filepath, as_attachment=True)


if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5050))
    debug = os.environ.get('FLASK_ENV') == 'development'
    app.run(host='0.0.0.0', port=port, debug=debug)
